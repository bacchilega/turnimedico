import datetime
import openpyxl
from openpyxl import Workbook
import random

# Funzione per verificare se un medico è disponibile per un determinato reparto
def is_doctor_available(doctor, department):
    if department in doctor_restrictions[doctor]:
        return False
    return True

# Funzione per assegnare un medico a un reparto per un giorno specifico
def assign_doctor_to_department(day, department):
    shuffled_doctors = doctors.copy()
    random.shuffle(shuffled_doctors)
    for doctor in shuffled_doctors:
        if is_doctor_available(doctor, department) and doctor not in assigned_doctors[day]:
            assigned_doctors[day].append(doctor)
            return doctor
    return None

# Inizializzazione dei dati
doctors = ['BANDINI', 'BEGHINI', 'BENINI', 'BISCIONE', 'CERÈ', 'DAMIANI', 'DE ROSA', 'DI MARIO', 'DI PIAZZA', 'FIORENTINI', 'FRAU', 'KRAUS', 'LABOMBARDA', 'MACCOLINI', 'MALFERRARI', 'MARCHETTI', 'PENAZZI', 'PERONI', 'PERRI', 'POLI', 'REMONDINI', 'SOFLAI', 'TASSINARI', 'TERRANOVA', 'ZANARDI']

doctor_restrictions = {
    'BANDINI': ['RMN'],
    'BEGHINI': ['RMN'],
    'BENINI': ['RMN'],
    'BISCIONE': ['GUARDIA_RIA', 'NOTTE_RIA', 'URG_POM', 'G_SEMI_P', 'REP', 'NOTTE_SALA', 'SALA1', 'SALA2', 'SALA3', 'SALA4', 'SALA5', 'SALA6', 'SALA_POM_1', 'SALA_POM_2', 'SALA_POM_3', 'SALA_POM_4', 'GASTRO', 'AMBULATORIO'],
    'CERÈ': ['RMN', 'GUARDIA_RIA', 'NOTTE_RIA', 'NOTTE_SALA', 'REP'],
    'DAMIANI': ['RMN'],
    'DE ROSA': ['RMN'],
    'DI MARIO': ['RMN'],
    'DI PIAZZA': ['RMN'],
    'FIORENTINI': ['RMN'],
    'FRAU': ['RMN'],
    'KRAUS': ['RMN'],
    'LABOMBARDA': ['RMN'],
    'MACCOLINI': [],
    'MALFERRARI': ['RMN'],
    'MARCHETTI': [],
    'PENAZZI': ['RMN'],
    'PERONI': ['RMN', 'GUARDIA_RIA', 'NOTTE_RIA', 'NOTTE_SALA', 'REP'],
    'PERRI': ['RMN'],
    'POLI': ['GUARDIA_RIA', 'NOTTE_RIA', 'URG_POM', 'G_SEMI_P', 'REP', 'NOTTE_SALA', 'SALA1', 'SALA2', 'SALA3', 'SALA4', 'SALA5', 'SALA6', 'SALA_POM_1', 'SALA_POM_2', 'SALA_POM_3', 'SALA_POM_4', 'GASTRO', 'AMBULATORIO'],
    'REMONDINI': ['RMN'],
    'SOFLAI': ['RMN'],
    'TASSINARI': ['RMN', 'GUARDIA_RIA', 'NOTTE_RIA', 'NOTTE_SALA', 'REP'],
    'TERRANOVA': ['RMN'],
    'ZANARDI': ['RMN']
}

departments = ['GUARDIA_RIA', 'NOTTE_RIA', 'URG_MATT', 'URG_POM', 'G_SEMI_M', 'G_SEMI_P', 'REP', 'NOTTE_SALA', 'SALA1', 'SALA2', 'SALA3', 'SALA4', 'SALA5', 'SALA6', 'SALA_POM_1', 'SALA_POM_2', 'SALA_POM_3', 'SALA_POM_4', 'RMN', 'GASTRO', 'AMBULATORIO']

start_date = datetime.date(2023, 4, 1)
end_date = datetime.date(2023, 4, 7)
days = (end_date - start_date).days + 1
assigned_doctors = {day: [] for day in range(days)}

# Assegna i medici ai reparti per ogni giorno
schedule = {}
for department in departments:
    schedule[department] = []
    for day in range(days):
        assigned_doctor = assign_doctor_to_department(day, department)
        schedule[department].append(assigned_doctor)

# Stampa il programma
print("REPARTO", end="")
for day in range(days):
    current_date = start_date + datetime.timedelta(days=day)
    print(f" | {current_date.strftime('%d %b (%a)').upper()}", end="")
print()

for department in departments:
    print(department, end="")
    for day in range(days):
        assigned_doctor = schedule[department][day]
        if assigned_doctor is not None:
            print(f" | {assigned_doctor}", end="")
        else:
            print(" | ", end="")
    print()

# Crea un nuovo foglio di calcolo di Excel
workbook = Workbook()
worksheet = workbook.active

# Scrivi l'intestazione delle colonne
worksheet.cell(row=1, column=1).value = "REPARTO"
for day in range(days):
    current_date = start_date + datetime.timedelta(days=day)
    worksheet.cell(row=1, column=day + 2).value = current_date.strftime('%d %b (%a)').upper()

# Scrivi i dati del programma nel foglio di calcolo
row_num = 2
for department in departments:
    worksheet.cell(row=row_num, column=1).value = department
    for day in range(days):
        assigned_doctor = schedule[department][day]
        if assigned_doctor is not None:
            worksheet.cell(row=row_num, column=day + 2).value = assigned_doctor
    row_num += 1

# Salva il foglio di calcolo in un file
workbook.save('schedule.xlsx')